import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os

url = 'https://biznes.interia.pl/gieldy/notowania-gpw#:~:text=Notowania%20Gie%C5%82dy%20Papier%C3%B3w%20Warto%C5%9Bciowych%20to,sprzeda%C5%BCy%20oraz%20kurs%20zawarcia%20transakcji.'
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')
titles = soup.select('p')
linki = soup.select('td')
list_of_comp = []
list_of_links = []
file_exel = 'Template.xlsx'
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "Name"
sheet["B1"] = "Price"
sheet["C1"] = "Change"
sheet["D1"] = "Open"
sheet["E1"] = "MAX"
sheet["F1"] = "MIN"
sheet["G1"] = "Ref"
sheet["H1"] = "Turnover"
sheet["I1"] = "Time"

if os.path.exists(file_exel):
    workbook = load_workbook(filename=file_exel)
else:
    workbook.save(filename=file_exel)

def for_nazwa():
    cleaned_list_nazwa = []
    for title in titles:
        remove_nec = title.text.replace(' ', '').replace('\xa0', '').replace('\n', '')
        cleaned_list_nazwa.append(remove_nec)
    return cleaned_list_nazwa

def for_link():
    cleaned_list_link = []
    for link in linki:
        remove_nec = link.text.replace(' ', '').replace('\xa0', '').replace('\n', '')
        if remove_nec == "" or remove_nec == " " or remove_nec == "Reklama":
            pass
        else:
            cleaned_list_link.append(remove_nec)
    return cleaned_list_link

my_cleaned_list_nazwa = for_nazwa()
my_cleaned_list_link = for_link()

print(my_cleaned_list_nazwa)
print(my_cleaned_list_link)

my_cleaned_list_nazwa.pop(0)
my_cleaned_list_nazwa.pop()

class Dane:
    def __init__(self, name, price, change, open, max, min, ref, turnover, time):
        self.all = name, price, change, open, max, min, ref, turnover, time

my_cleaned_list_dane = []
for i in range(0, len(my_cleaned_list_link), 9):
    sublist = my_cleaned_list_link[i:i+9]
    if len(sublist) == 9:
        my_cleaned_list_dane.append(Dane(*sublist))

for dane in my_cleaned_list_dane:
    sheet.append(dane.all)

workbook.save(filename=file_exel)
